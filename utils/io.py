import json
from utils.formatter import format_lead
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

LEADS_FILE = "leads.json"

def load_leads():
    try:
        with open(LEADS_FILE, "r") as file:
            return json.load(file)
    except FileNotFoundError:
        return []

def save_leads(leads):
    with open(LEADS_FILE, "w") as file:
        json.dump(leads, file, indent=2)

def export_formatted(leads, filename="formatted_leads.md"):
    with open(filename, "w") as file:
        for lead in leads:
            file.write(format_lead(lead))
            file.write("\n")

def export_to_csv(leads, filename="leads_export.csv"):
    df = pd.DataFrame(leads)
    df_view = df[[
        "company_name", "company_website", "company_linkedin",
        "industry", "product_services", "hq_city", "hq_state",
        "year_founded", "ceo_name", "source",
        "operational_fit", "ethical_alignment", "growth_signals",
        "research_notes"
    ]].rename(columns={
        "company_name": "Company",
        "company_website": "Website",
        "company_linkedin": "LinkedIn",
        "industry": "Industry",
        "product_services": "Product/Services",
        "hq_city": "City",
        "hq_state": "State",
        "year_founded": "Founded",
        "ceo_name": "CEO",
        "source": "Source",
        "operational_fit": "Operational Fit",
        "ethical_alignment": "Ethical Alignment",
        "growth_signals": "Growth Signals",
        "research_notes": "Research Notes/Context"
    })
    df_view.to_csv(filename, index=False)
    print(f"📤 Leads exported to CSV: {filename}")

def export_to_excel(leads, filename="leads_export.xlsx"):
    rows = []
    for lead in leads:
        rows.append({
            "Company": lead.get("company_name", ""),
            "Website": lead.get("company_website", ""),
            "LinkedIn": lead.get("company_linkedin", ""),
            "Industry": lead.get("industry", ""),
            "Product/Services": lead.get("product_services", ""),
            "City": lead.get("hq_city", ""),
            "State": lead.get("hq_state", ""),
            "Founded": lead.get("year_founded", ""),
            "CEO": lead.get("ceo_name", ""),
            "Source": lead.get("source", ""),
            "Operational Fit": lead.get("operational_fit", ""),
            "Ethical Alignment": lead.get("ethical_alignment", ""),
            "Growth Signals": lead.get("growth_signals", ""),
            "Research Notes/Context": lead.get("research_notes", "")
        })

    df = pd.DataFrame(rows)
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(filename)
    print(f"📊 Excel file styled and saved as: {filename}")
